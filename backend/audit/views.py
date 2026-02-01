"""
Views for the REdI Trolley Audit System.

Implements all views for the audit workflow:
- Dashboard with summary statistics
- Trolley/location management (list, detail, edit)
- Multi-step audit wizard (documents, equipment, condition, checks, review, submit)
- Issue lifecycle (list, create, detail, transitions, comments)
- Random audit selection (view, generate)
- Reports and CSV export
"""

import logging
from datetime import date, timedelta

import csv

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Avg, Count, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView, DetailView, ListView, TemplateView, UpdateView,
)

from .forms import (
    AuditChecksForm, AuditConditionForm, AuditDocumentsForm,
    CorrectiveActionForm, IssueAssignForm, IssueCommentForm,
    IssueCreateForm, IssueEditForm, IssueResolveForm, LocationEditForm,
)
from .mixins import (
    AuditorRequiredMixin, EducatorRequiredMixin, ManagerRequiredMixin,
    ViewerRequiredMixin,
)
from .models import (
    Audit, AuditChecks, AuditCondition, AuditDocuments, AuditEquipment,
    AuditPeriod, CorrectiveAction, Equipment, Issue, Location,
    LocationEquipment, RandomAuditSelection, RandomAuditSelectionItem,
    ServiceLine,
)
from .services.compliance import ComplianceScorer
from .services.issue_workflow import InvalidTransitionError, IssueWorkflow
from .services.notifications import NotificationService
from .services.random_selection import RandomAuditSelector

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Mixins for Audit Wizard Ownership Validation
# ---------------------------------------------------------------------------

class AuditOwnershipMixin:
    """
    Mixin to validate that the current user owns the audit and it's in progress.
    Used by all audit wizard views to enforce ownership checks.
    """
    def dispatch(self, request, *args, **kwargs):
        # Get the audit object
        audit = get_object_or_404(Audit, pk=kwargs.get('pk'))

        # Check ownership: user must be the auditor or a superuser
        if audit.auditor_user != request.user and not request.user.is_superuser:
            messages.error(
                request,
                'You do not have permission to edit this audit. Only the original auditor can modify it.',
            )
            return HttpResponseForbidden(
                'You do not have permission to edit this audit.',
            )

        # Check status: audit must be in progress
        if audit.submission_status != 'InProgress':
            messages.error(
                request,
                'This audit has already been submitted and cannot be modified.',
            )
            return redirect('audit:audit_detail', pk=audit.pk)

        return super().dispatch(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

class DashboardView(LoginRequiredMixin, TemplateView):
    """Main dashboard with summary statistics and recent activity."""

    template_name = 'audit/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total_locations'] = Location.objects.filter(status='Active').count()
        ctx['total_audits'] = Audit.objects.filter(
            submission_status='Submitted',
        ).count()
        ctx['open_issues'] = Issue.objects.exclude(
            status__in=['Closed', 'Resolved'],
        ).count()
        ctx['avg_compliance'] = Audit.objects.filter(
            submission_status='Submitted',
            overall_compliance__isnull=False,
        ).aggregate(avg=Avg('overall_compliance'))['avg']

        # Recent audits
        ctx['recent_audits'] = (
            Audit.objects.filter(submission_status='Submitted')
            .select_related('location', 'period')
            .order_by('-completed_at')[:10]
        )

        # Recent issues
        ctx['recent_issues'] = (
            Issue.objects.select_related('location', 'equipment')
            .order_by('-reported_date')[:10]
        )

        # Active random selection
        selector = RandomAuditSelector()
        ctx['active_selection'] = selector.get_active_selection()

        # Compliance by service line
        ctx['service_line_stats'] = (
            ServiceLine.objects.filter(is_active=True)
            .annotate(
                location_count=Count(
                    'locations',
                    filter=Q(locations__status='Active'),
                ),
                audit_count=Count(
                    'locations__audits',
                    filter=Q(locations__audits__submission_status='Submitted'),
                ),
                avg_compliance=Avg(
                    'locations__audits__overall_compliance',
                    filter=Q(locations__audits__submission_status='Submitted'),
                ),
            )
            .order_by('name')
        )

        return ctx


# ---------------------------------------------------------------------------
# Trolley / Location views
# ---------------------------------------------------------------------------

class TrolleyListView(ViewerRequiredMixin, ListView):
    """Paginated, filterable list of trolley locations."""

    model = Location
    template_name = 'audit/trolley_list.html'
    context_object_name = 'trolleys'
    paginate_by = 25

    def get_queryset(self):
        qs = Location.objects.select_related('service_line').order_by(
            'department_name',
        )

        service_line = self.request.GET.get('service_line')
        if service_line:
            qs = qs.filter(service_line_id=service_line)

        status = self.request.GET.get('status', 'Active')
        if status:
            qs = qs.filter(status=status)

        building = self.request.GET.get('building')
        if building:
            qs = qs.filter(building=building)

        search = self.request.GET.get('q')
        if search:
            qs = qs.filter(
                Q(department_name__icontains=search)
                | Q(display_name__icontains=search)
                | Q(building__icontains=search)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['service_lines'] = ServiceLine.objects.filter(is_active=True)
        ctx['buildings'] = (
            Location.objects.values_list('building', flat=True)
            .distinct()
            .order_by('building')
        )
        ctx['current_filters'] = {
            'service_line': self.request.GET.get('service_line', ''),
            'status': self.request.GET.get('status', 'Active'),
            'building': self.request.GET.get('building', ''),
            'q': self.request.GET.get('q', ''),
        }
        return ctx


class TrolleyDetailView(ViewerRequiredMixin, DetailView):
    """Detail page for a single trolley location with audit history."""

    model = Location
    template_name = 'audit/trolley_detail.html'
    context_object_name = 'trolley'

    def get_queryset(self):
        return Location.objects.select_related('service_line')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        trolley = self.object
        ctx['recent_audits'] = (
            trolley.audits.filter(submission_status='Submitted')
            .order_by('-completed_at')[:10]
        )
        ctx['open_issues'] = trolley.issues.exclude(
            status__in=['Closed', 'Resolved'],
        ).order_by('-reported_date')
        ctx['change_logs'] = trolley.change_logs.all()[:20]
        return ctx


class TrolleyEditView(EducatorRequiredMixin, UpdateView):
    """Edit form for trolley location details."""

    model = Location
    form_class = LocationEditForm
    template_name = 'audit/trolley_edit.html'

    def get_success_url(self):
        messages.success(self.request, 'Trolley updated successfully.')
        return self.object.get_absolute_url()


# ---------------------------------------------------------------------------
# Audit views
# ---------------------------------------------------------------------------

class AuditListView(ViewerRequiredMixin, ListView):
    """Paginated, filterable list of audits."""

    model = Audit
    template_name = 'audit/audit_list.html'
    context_object_name = 'audits'
    paginate_by = 25

    def get_queryset(self):
        qs = Audit.objects.select_related(
            'location', 'location__service_line', 'period',
        ).order_by('-started_at')

        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(submission_status=status)

        service_line = self.request.GET.get('service_line')
        if service_line:
            qs = qs.filter(location__service_line_id=service_line)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['service_lines'] = ServiceLine.objects.filter(is_active=True)
        ctx['current_filters'] = {
            'status': self.request.GET.get('status', ''),
            'service_line': self.request.GET.get('service_line', ''),
        }
        return ctx


class AuditStartView(AuditorRequiredMixin, View):
    """Start a new audit for a location (POST only)."""

    def post(self, request, location_id):
        location = get_object_or_404(Location, pk=location_id)

        # Prevent duplicate in-progress audits for the same location
        existing = Audit.objects.filter(
            location=location,
            submission_status='InProgress',
        ).first()
        if existing:
            messages.info(
                request,
                f'An audit is already in progress for {location.display_name}. Resuming existing audit.',
            )
            return redirect('audit:audit_documents', pk=existing.pk)

        period = AuditPeriod.objects.filter(is_active=True).first()

        if not period:
            messages.error(request, 'No active audit period found.')
            return redirect('audit:trolley_detail', pk=location_id)

        # Determine audit type based on context
        audit_type = 'Monthly'
        selection_item_id = request.POST.get('selection_item') or request.GET.get('selection_item')
        if selection_item_id:
            audit_type = 'Random'

        audit = Audit.objects.create(
            location=location,
            period=period,
            audit_type=audit_type,
            auditor_name=(
                request.user.get_full_name() or request.user.username
            ),
            auditor_user=request.user,
            submission_status='InProgress',
        )

        # Create related section records
        AuditDocuments.objects.create(audit=audit)
        AuditCondition.objects.create(audit=audit)
        AuditChecks.objects.create(
            audit=audit,
            expected_outside=(
                period.expected_outside_checks_24_7
                if location.operating_hours == '24_7'
                else 0
            ),
            expected_inside=period.expected_inside_checks,
        )

        # Create equipment check records filtered to this location
        equipment_list = Equipment.objects.filter(is_active=True)

        # Prefetch all equipment overrides for this location to avoid N+1 queries
        overrides = {
            le.equipment_id: le
            for le in LocationEquipment.objects.filter(location=location)
        }

        for equip in equipment_list:
            # Skip items not matching defibrillator type
            if (
                equip.required_for_defib_type != 'N/A'
                and equip.required_for_defib_type != location.defibrillator_type
            ):
                continue
            # Skip paediatric items when location has no paediatric box
            if equip.is_paediatric_item and not location.has_paediatric_box:
                continue
            # Skip altered airway items when location has no altered airway kit
            if equip.is_altered_airway_item and not location.has_altered_airway:
                continue

            # Check for custom quantity override using prefetched dict
            override = overrides.get(equip.pk)
            expected_qty = (
                override.custom_quantity
                if override and override.custom_quantity
                else equip.standard_quantity
            )

            AuditEquipment.objects.create(
                audit=audit,
                equipment=equip,
                quantity_expected=expected_qty,
            )

        messages.success(
            request, f'Audit started for {location.display_name}.',
        )
        return redirect('audit:audit_documents', pk=audit.pk)


class AuditDetailView(ViewerRequiredMixin, DetailView):
    """Read-only detail view for a completed or in-progress audit."""

    model = Audit
    template_name = 'audit/audit_detail.html'
    context_object_name = 'audit'

    def get_queryset(self):
        return Audit.objects.select_related(
            'location', 'location__service_line', 'period',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        audit = self.object
        try:
            ctx['documents'] = audit.documents
        except AuditDocuments.DoesNotExist:
            ctx['documents'] = None
        try:
            ctx['condition'] = audit.condition
        except AuditCondition.DoesNotExist:
            ctx['condition'] = None
        try:
            ctx['checks'] = audit.checks
        except AuditChecks.DoesNotExist:
            ctx['checks'] = None
        ctx['equipment_checks'] = (
            audit.equipment_checks
            .select_related('equipment', 'equipment__category')
            .order_by('equipment__category__sort_order', 'equipment__sort_order')
        )
        ctx['issues'] = audit.issues.all()

        # Check if user is educator for follow-up audit button
        ctx['is_educator'] = (
            self.request.user.groups.filter(name__in=['MERT Educator', 'System Admin']).exists()
            or self.request.user.is_superuser
        )

        return ctx


class AuditDocumentsView(AuditOwnershipMixin, AuditorRequiredMixin, View):
    """Step 1 of audit wizard: documentation checks."""

    template_name = 'audit/audit_documents.html'

    def get(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        documents, _ = AuditDocuments.objects.get_or_create(audit=audit)
        form = AuditDocumentsForm(instance=documents)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })

    def post(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        documents, _ = AuditDocuments.objects.get_or_create(audit=audit)
        form = AuditDocumentsForm(request.POST, instance=documents)
        if form.is_valid():
            form.save()
            messages.success(request, 'Documentation check saved.')
            return redirect('audit:audit_equipment', pk=audit.pk)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })


class AuditEquipmentView(AuditOwnershipMixin, AuditorRequiredMixin, View):
    """Step 2 of audit wizard: equipment checklist."""

    template_name = 'audit/audit_equipment.html'

    def get(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        equipment_checks = (
            audit.equipment_checks
            .select_related('equipment', 'equipment__category')
            .order_by('equipment__category__sort_order', 'equipment__sort_order')
        )

        # Group by category for template rendering
        categories = {}
        for check in equipment_checks:
            cat_name = check.equipment.category.category_name
            if cat_name not in categories:
                categories[cat_name] = []
            categories[cat_name].append(check)

        return render(request, self.template_name, {
            'audit': audit, 'categories': categories,
        })

    def post(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        equipment_checks = audit.equipment_checks.all()

        try:
            with transaction.atomic():
                for check in equipment_checks:
                    prefix = f'equip_{check.pk}'
                    check.is_present = request.POST.get(f'{prefix}_present') == 'on'
                    qty_value = request.POST.get(f'{prefix}_qty', '0')
                    try:
                        check.quantity_found = int(qty_value)
                    except ValueError:
                        messages.error(
                            request,
                            f'Invalid quantity value for {check.equipment.item_name}',
                        )
                        return redirect('audit:audit_equipment', pk=audit.pk)
                    check.expiry_ok = request.POST.get(f'{prefix}_expiry') == 'on'
                    check.item_notes = request.POST.get(f'{prefix}_notes', '')
                    check.save()
        except Exception as e:
            messages.error(request, f'Error saving equipment checklist: {e}')
            return redirect('audit:audit_equipment', pk=audit.pk)

        messages.success(request, 'Equipment checklist saved.')
        return redirect('audit:audit_condition', pk=audit.pk)


class AuditConditionView(AuditOwnershipMixin, AuditorRequiredMixin, View):
    """Step 3 of audit wizard: physical condition assessment."""

    template_name = 'audit/audit_condition.html'

    def get(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        condition, _ = AuditCondition.objects.get_or_create(audit=audit)
        form = AuditConditionForm(instance=condition)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })

    def post(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        condition, _ = AuditCondition.objects.get_or_create(audit=audit)
        form = AuditConditionForm(request.POST, instance=condition)
        if form.is_valid():
            form.save()
            messages.success(request, 'Condition check saved.')
            return redirect('audit:audit_checks', pk=audit.pk)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })


class AuditChecksView(AuditOwnershipMixin, AuditorRequiredMixin, View):
    """Step 4 of audit wizard: routine check counts."""

    template_name = 'audit/audit_checks.html'

    def get(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        checks, _ = AuditChecks.objects.get_or_create(audit=audit)
        form = AuditChecksForm(instance=checks)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })

    def post(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)
        checks, _ = AuditChecks.objects.get_or_create(audit=audit)
        form = AuditChecksForm(request.POST, instance=checks)
        if form.is_valid():
            form.save()
            messages.success(request, 'Routine checks saved.')
            return redirect('audit:audit_review', pk=audit.pk)
        return render(request, self.template_name, {
            'audit': audit, 'form': form,
        })


class AuditReviewView(AuditOwnershipMixin, AuditorRequiredMixin, DetailView):
    """Step 5 of audit wizard: review all sections before submission."""

    model = Audit
    template_name = 'audit/audit_review.html'
    context_object_name = 'audit'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        audit = self.object

        try:
            ctx['documents'] = audit.documents
        except AuditDocuments.DoesNotExist:
            ctx['documents'] = None
        try:
            ctx['condition'] = audit.condition
        except AuditCondition.DoesNotExist:
            ctx['condition'] = None
        try:
            ctx['checks'] = audit.checks
        except AuditChecks.DoesNotExist:
            ctx['checks'] = None

        ctx['equipment_checks'] = (
            audit.equipment_checks
            .select_related('equipment', 'equipment__category')
            .order_by('equipment__category__sort_order', 'equipment__sort_order')
        )

        # Calculate score previews
        scorer = ComplianceScorer()
        try:
            ctx['doc_score'] = scorer.calculate_documentation_score(
                audit.documents,
            )
        except Exception:
            ctx['doc_score'] = None
        ctx['equip_score'] = scorer.calculate_equipment_score(
            audit.equipment_checks.all(),
        )
        try:
            ctx['cond_score'] = scorer.calculate_condition_score(
                audit.condition,
            )
        except Exception:
            ctx['cond_score'] = None
        try:
            ctx['check_score'] = scorer.calculate_check_score(audit.checks)
        except Exception:
            ctx['check_score'] = None

        return ctx


class AuditSubmitView(AuditOwnershipMixin, AuditorRequiredMixin, View):
    """Submit a completed audit and calculate compliance scores (POST only)."""

    def post(self, request, pk):
        audit = get_object_or_404(Audit, pk=pk)

        # Calculate compliance scores
        scorer = ComplianceScorer()
        overall = scorer.calculate_overall_score(audit)

        # Wrap all database operations in a transaction
        with transaction.atomic():
            # Update audit status
            audit.submission_status = 'Submitted'
            audit.completed_at = timezone.now()
            audit.save(update_fields=['submission_status', 'completed_at'])

            # Update location's last audit info
            location = audit.location
            location.last_audit_date = audit.completed_at
            location.last_audit_compliance = overall
            location.save(update_fields=['last_audit_date', 'last_audit_compliance'])

            # Mark any related random selection item as completed
            from .models import RandomAuditSelectionItem
            RandomAuditSelectionItem.objects.filter(
                location=location,
                audit_status='Pending',
                selection__is_active=True,
            ).update(
                audit_status='Completed',
                audit=audit,
            )

            # Auto-create issues for critical equipment failures
            critical_issues = []
            for check in audit.equipment_checks.select_related('equipment'):
                if check.equipment.critical_item and not check.is_present:
                    issue = Issue.objects.create(
                        location=location,
                        audit=audit,
                        issue_category='Equipment',
                        severity='Critical',
                        title=f'Missing critical item: {check.equipment.item_name}',
                        description=(
                            f'Critical equipment item '
                            f'"{check.equipment.item_name}" was not found '
                            f'during audit.'
                        ),
                        equipment=check.equipment,
                        reported_by=audit.auditor_name,
                    )
                    critical_issues.append(issue)

            # Low compliance triggers follow-up
            if overall < 80:
                audit.requires_follow_up = True
                audit.follow_up_due_date = date.today() + timedelta(days=14)
                audit.save(
                    update_fields=['requires_follow_up', 'follow_up_due_date'],
                )

        # Send notifications (never crash on email failure)
        try:
            notifications = NotificationService()
            notifications.notify_audit_completed(audit)
            for issue in critical_issues:
                notifications.notify_critical_issue(issue)
        except Exception:
            logger.warning(
                'Failed to send notifications for audit %s',
                audit.pk,
                exc_info=True,
            )

        messages.success(
            request,
            f'Audit submitted. Overall compliance: {overall}%',
        )
        return redirect('audit:audit_detail', pk=audit.pk)


# ---------------------------------------------------------------------------
# Issue views
# ---------------------------------------------------------------------------

class IssueListView(ViewerRequiredMixin, ListView):
    """Paginated, filterable list of issues with status counts."""

    model = Issue
    template_name = 'audit/issue_list.html'
    context_object_name = 'issues'
    paginate_by = 25

    def get_queryset(self):
        qs = Issue.objects.select_related(
            'location', 'location__service_line', 'equipment',
        ).order_by('-reported_date')

        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)

        severity = self.request.GET.get('severity')
        if severity:
            qs = qs.filter(severity=severity)

        service_line = self.request.GET.get('service_line')
        if service_line:
            qs = qs.filter(location__service_line_id=service_line)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['service_lines'] = ServiceLine.objects.filter(is_active=True)
        ctx['current_filters'] = {
            'status': self.request.GET.get('status', ''),
            'severity': self.request.GET.get('severity', ''),
            'service_line': self.request.GET.get('service_line', ''),
        }
        # Issue counts by status - use single aggregate query to avoid N+1
        status_counts = Issue.objects.aggregate(
            **{
                status_code: Count('id', filter=Q(status=status_code))
                for status_code, _status_label in Issue.STATUS_CHOICES
            }
        )
        ctx['status_counts'] = status_counts
        return ctx


class IssueCreateView(AuditorRequiredMixin, CreateView):
    """Create a new issue with automatic SLA target date."""

    model = Issue
    form_class = IssueCreateForm
    template_name = 'audit/issue_create.html'

    def form_valid(self, form):
        form.instance.reported_by = (
            self.request.user.get_full_name() or self.request.user.username
        )
        response = super().form_valid(form)

        # Set SLA target date based on severity
        workflow = IssueWorkflow()
        workflow.set_target_resolution_date(self.object)

        # Notify if critical issue
        notifications = NotificationService()
        notifications.notify_critical_issue(self.object)

        messages.success(
            self.request,
            f'Issue {self.object.issue_number} created.',
        )
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()


class IssueDetailView(ViewerRequiredMixin, DetailView):
    """Full issue detail with workflow actions, comments, and SLA status."""

    model = Issue
    template_name = 'audit/issue_detail.html'
    context_object_name = 'issue'

    def get_queryset(self):
        return Issue.objects.select_related(
            'location', 'location__service_line', 'equipment', 'audit',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        issue = self.object
        ctx['corrective_actions'] = issue.corrective_actions.all()
        ctx['comments'] = issue.comments.all()
        ctx['comment_form'] = IssueCommentForm()
        ctx['assign_form'] = IssueAssignForm()
        ctx['resolve_form'] = IssueResolveForm()

        workflow = IssueWorkflow()
        ctx['available_transitions'] = workflow.get_available_transitions(issue)
        ctx['sla_target'] = workflow.get_sla_target(issue)
        ctx['sla_breached'] = workflow.is_sla_breached(issue)
        ctx['needs_management_review'] = workflow.needs_management_review(
            issue,
        )

        return ctx


class IssueTransitionView(ManagerRequiredMixin, View):
    """Handle issue state transitions (POST only)."""

    def post(self, request, pk, action):
        issue = get_object_or_404(Issue, pk=pk)
        workflow = IssueWorkflow()
        user_name = (
            request.user.get_full_name() or request.user.username
        )

        try:
            if action == 'assign':
                assigned_to = request.POST.get('assigned_to', '')
                workflow.assign(issue, assigned_to, user_name)
                notifications = NotificationService()
                notifications.notify_issue_assigned(issue)
                messages.success(
                    request, f'Issue assigned to {assigned_to}.',
                )
            elif action == 'start':
                workflow.start_work(issue, user_name)
                messages.success(request, 'Work started on issue.')
            elif action == 'submit_verification':
                summary = request.POST.get('resolution_summary', '')
                workflow.submit_for_verification(issue, user_name, summary)
                messages.success(request, 'Submitted for verification.')
            elif action == 'verify':
                workflow.verify_and_resolve(issue, user_name)
                messages.success(request, 'Issue verified and resolved.')
            elif action == 'reject':
                reason = request.POST.get('reason', '')
                workflow.reject_verification(issue, user_name, reason)
                messages.success(request, 'Verification rejected.')
            elif action == 'close':
                workflow.close(issue, user_name)
                messages.success(request, 'Issue closed.')
            elif action == 'reopen':
                reason = request.POST.get('reason', '')
                workflow.reopen(issue, user_name, reason)
                messages.success(request, 'Issue reopened.')
            elif action == 'escalate':
                reason = request.POST.get('reason', '')
                workflow.escalate(issue, user_name, reason)
                messages.success(request, 'Issue escalated.')
            else:
                messages.error(request, f'Unknown action: {action}')
        except InvalidTransitionError as e:
            messages.error(request, str(e))

        return redirect('audit:issue_detail', pk=issue.pk)


class IssueCommentView(AuditorRequiredMixin, View):
    """Add a comment to an issue (POST only)."""

    def post(self, request, pk):
        issue = get_object_or_404(Issue, pk=pk)
        form = IssueCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.issue = issue
            comment.comment_by = (
                request.user.get_full_name() or request.user.username
            )
            comment.save()
            messages.success(request, 'Comment added.')
        return redirect('audit:issue_detail', pk=issue.pk)


class CorrectiveActionCreateView(ManagerRequiredMixin, View):
    """Record a corrective action for an issue (POST only)."""

    def post(self, request, pk):
        issue = get_object_or_404(Issue, pk=pk)
        form = CorrectiveActionForm(request.POST)
        if form.is_valid():
            action = form.save(commit=False)
            action.issue = issue
            action.action_taken_by = (
                request.user.get_full_name() or request.user.username
            )
            # Auto-assign action number
            last_action = issue.corrective_actions.order_by('-action_number').first()
            action.action_number = (last_action.action_number + 1) if last_action else 1
            action.save()
            messages.success(request, 'Corrective action recorded.')
        else:
            messages.error(request, 'Please correct the form errors.')
        return redirect('audit:issue_detail', pk=issue.pk)


class IssueEditView(ManagerRequiredMixin, UpdateView):
    """Edit an existing issue's title, description, severity, or category."""

    model = Issue
    form_class = IssueEditForm
    template_name = 'audit/issue_edit.html'

    def get_success_url(self):
        messages.success(self.request, 'Issue updated successfully.')
        return self.object.get_absolute_url()


# ---------------------------------------------------------------------------
# Random Selection views
# ---------------------------------------------------------------------------

class RandomSelectionView(EducatorRequiredMixin, TemplateView):
    """View current and past weekly random audit selections."""

    template_name = 'audit/random_selection.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        selector = RandomAuditSelector()
        ctx['active_selection'] = selector.get_active_selection()
        ctx['past_selections'] = (
            RandomAuditSelection.objects.filter(is_active=False)
            .order_by('-week_start_date')[:10]
        )
        return ctx


class GenerateSelectionView(EducatorRequiredMixin, View):
    """Generate a new weekly random audit selection (POST only)."""

    def post(self, request):
        selector = RandomAuditSelector()
        user_name = (
            request.user.get_full_name() or request.user.username
        )
        selection = selector.generate_selection(generated_by=user_name)
        messages.success(
            request,
            f'Generated selection with {selection.items.count()} trolleys.',
        )
        return redirect('audit:random_selection')


class SelectionDetailView(EducatorRequiredMixin, DetailView):
    """Detail view for a single random audit selection with item status."""

    model = RandomAuditSelection
    template_name = 'audit/selection_detail.html'
    context_object_name = 'selection'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = (
            self.object.items
            .select_related('location', 'location__service_line', 'audit')
            .order_by('selection_rank')
        )
        total = self.object.items.count()
        completed = self.object.items.filter(audit_status='Completed').count()
        ctx['completion_pct'] = int(completed / total * 100) if total > 0 else 0
        ctx['completed_count'] = completed
        ctx['total_count'] = total
        return ctx


# ---------------------------------------------------------------------------
# Reports views
# ---------------------------------------------------------------------------

class ReportsView(ViewerRequiredMixin, TemplateView):
    """Reports landing page with summary metrics."""

    template_name = 'audit/reports.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total_audits'] = Audit.objects.filter(
            submission_status='Submitted',
        ).count()
        ctx['total_issues'] = Issue.objects.count()
        ctx['avg_compliance'] = Audit.objects.filter(
            submission_status='Submitted',
            overall_compliance__isnull=False,
        ).aggregate(avg=Avg('overall_compliance'))['avg']
        ctx['service_lines'] = ServiceLine.objects.filter(is_active=True)
        return ctx


class ComplianceReportView(ViewerRequiredMixin, TemplateView):
    """Detailed compliance report by service line."""

    template_name = 'audit/compliance_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx['service_line_compliance'] = (
            ServiceLine.objects.filter(is_active=True)
            .annotate(
                avg_compliance=Avg(
                    'locations__audits__overall_compliance',
                    filter=Q(
                        locations__audits__submission_status='Submitted',
                    ),
                ),
                audit_count=Count(
                    'locations__audits',
                    filter=Q(
                        locations__audits__submission_status='Submitted',
                    ),
                ),
                location_count=Count(
                    'locations',
                    filter=Q(locations__status='Active'),
                ),
            )
            .order_by('name')
        )

        ctx['recent_audits'] = (
            Audit.objects.filter(submission_status='Submitted')
            .select_related('location', 'location__service_line')
            .order_by('-completed_at')[:50]
        )

        return ctx


class ComplianceTrendApiView(ViewerRequiredMixin, View):
    """JSON API: monthly average compliance over the last 12 months."""

    def get(self, request):
        twelve_months_ago = timezone.now() - timedelta(days=365)
        data = (
            Audit.objects.filter(
                submission_status='Submitted',
                completed_at__gte=twelve_months_ago,
            )
            .annotate(month=TruncMonth('completed_at'))
            .values('month')
            .annotate(avg_compliance=Avg('overall_compliance'))
            .order_by('month')
        )

        labels = []
        values = []
        for entry in data:
            labels.append(entry['month'].strftime('%b %Y'))
            values.append(float(entry['avg_compliance'] or 0))

        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'label': 'Average Compliance %',
                'data': values,
            }],
        })


class IssuesBySeverityApiView(ViewerRequiredMixin, View):
    """JSON API: open issues grouped by severity."""

    def get(self, request):
        data = (
            Issue.objects.exclude(status__in=['Closed', 'Resolved'])
            .values('severity')
            .annotate(count=Count('id'))
            .order_by('severity')
        )

        labels = []
        values = []
        colors = {
            'Critical': '#E55B64',
            'High': '#FFC107',
            'Medium': '#FF9800',
            'Low': '#6C757D',
        }
        bg_colors = []

        for entry in data:
            labels.append(entry['severity'])
            values.append(entry['count'])
            bg_colors.append(colors.get(entry['severity'], '#6C757D'))

        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'data': values,
                'backgroundColor': bg_colors,
            }],
        })


class AuditVolumeApiView(ViewerRequiredMixin, View):
    """JSON API: audits per month over the last 12 months."""

    def get(self, request):
        twelve_months_ago = timezone.now() - timedelta(days=365)
        data = (
            Audit.objects.filter(
                submission_status='Submitted',
                completed_at__gte=twelve_months_ago,
            )
            .annotate(month=TruncMonth('completed_at'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )

        labels = []
        values = []
        for entry in data:
            labels.append(entry['month'].strftime('%b %Y'))
            values.append(entry['count'])

        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'label': 'Audits Completed',
                'data': values,
            }],
        })


class ExportView(ManagerRequiredMixin, View):
    """Export audit data as CSV or Excel. Supports audits, issues, and locations."""

    ALLOWED_EXPORT_TYPES = {'audits', 'issues', 'locations'}

    def get(self, request):
        export_type = request.GET.get('type', 'audits')
        export_format = request.GET.get('format', 'csv')

        # Validate export type against whitelist
        if export_type not in self.ALLOWED_EXPORT_TYPES:
            return HttpResponseBadRequest(
                f'Invalid export type. Must be one of: {", ".join(self.ALLOWED_EXPORT_TYPES)}',
            )

        if export_format == 'xlsx':
            return self._export_xlsx(export_type)
        return self._export_csv(export_type)

    def _export_csv(self, export_type):
        """Export data as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="redi_{export_type}_{date.today()}.csv"'
        )
        writer = csv.writer(response)

        if export_type == 'audits':
            self._write_audit_rows(writer)
        elif export_type == 'issues':
            self._write_issue_rows(writer)
        elif export_type == 'locations':
            self._write_location_rows(writer)

        return response

    def _export_xlsx(self, export_type):
        """Export data as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return HttpResponseBadRequest('Excel export requires openpyxl.')

        wb = Workbook()
        ws = wb.active
        ws.title = export_type.capitalize()

        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1B3A5F', end_color='1B3A5F', fill_type='solid')

        class XlsxWriter:
            """Adapter to write rows to openpyxl worksheet."""
            def __init__(self, worksheet):
                self.ws = worksheet
                self.row_num = 0

            def writerow(self, row):
                self.row_num += 1
                for col_num, value in enumerate(row, 1):
                    cell = self.ws.cell(row=self.row_num, column=col_num, value=value)
                    if self.row_num == 1:
                        cell.font = header_font
                        cell.fill = header_fill

        writer = XlsxWriter(ws)

        if export_type == 'audits':
            self._write_audit_rows(writer)
        elif export_type == 'issues':
            self._write_issue_rows(writer)
        elif export_type == 'locations':
            self._write_location_rows(writer)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="redi_{export_type}_{date.today()}.xlsx"'
        )
        return response

    def _write_audit_rows(self, writer):
        writer.writerow([
            'Location', 'Service Line', 'Audit Date', 'Auditor',
            'Overall %', 'Documentation %', 'Equipment %',
            'Condition %', 'Checks %', 'Status',
        ])
        audits = (
            Audit.objects.filter(submission_status='Submitted')
            .select_related('location', 'location__service_line')
        )
        for audit in audits:
            writer.writerow([
                audit.location.display_name,
                audit.location.service_line.abbreviation,
                (
                    audit.completed_at.strftime('%Y-%m-%d')
                    if audit.completed_at else ''
                ),
                audit.auditor_name,
                float(audit.overall_compliance) if audit.overall_compliance else '',
                float(audit.document_score) if audit.document_score else '',
                float(audit.equipment_score) if audit.equipment_score else '',
                float(audit.condition_score) if audit.condition_score else '',
                float(audit.check_score) if audit.check_score else '',
                audit.submission_status,
            ])

    def _write_issue_rows(self, writer):
        writer.writerow([
            'Issue #', 'Location', 'Category', 'Severity',
            'Title', 'Status', 'Reported Date', 'Assigned To',
            'Escalation Level',
        ])
        for issue in Issue.objects.select_related('location'):
            writer.writerow([
                issue.issue_number,
                issue.location.display_name,
                issue.issue_category,
                issue.severity,
                issue.title,
                issue.status,
                (
                    issue.reported_date.strftime('%Y-%m-%d')
                    if issue.reported_date else ''
                ),
                issue.assigned_to,
                issue.escalation_level,
            ])

    def _write_location_rows(self, writer):
        writer.writerow([
            'Department', 'Display Name', 'Service Line', 'Building',
            'Level', 'Operating Hours', 'Has Paed Box', 'Defib Type',
            'Last Audit', 'Last Compliance %', 'Status',
        ])
        for loc in Location.objects.select_related('service_line'):
            writer.writerow([
                loc.department_name,
                loc.display_name,
                loc.service_line.abbreviation,
                loc.building,
                loc.level,
                loc.get_operating_hours_display(),
                'Yes' if loc.has_paediatric_box else 'No',
                loc.get_defibrillator_type_display(),
                (
                    loc.last_audit_date.strftime('%Y-%m-%d')
                    if loc.last_audit_date else 'Never'
                ),
                float(loc.last_audit_compliance) if loc.last_audit_compliance else 'N/A',
                loc.status,
            ])
